# This files contains your custom actions which can be used to run
# custom Python code.
#
# See this guide on how to implement these action:
# https://rasa.com/docs/rasa/custom-actions


# This is a simple example for a custom action which utters "Hello World!"

from typing import Any, Text, Dict, List

from rasa_sdk import Action, Tracker
from rasa_sdk.executor import CollectingDispatcher
import openpyxl
#
#
# class ActionHelloWorld(Action):
#
#     def name(self) -> Text:
#         return "action_hello_world"
#
#     def run(self, dispatcher: CollectingDispatcher,
#             tracker: Tracker,
#             domain: Dict[Text, Any]) -> List[Dict[Text, Any]]:
#
#         dispatcher.utter_message(text="Hello World!")
#
#         return []
class ActionChiTieuTuyenSinh(Action):

    def name(self) -> Text:
        return "action_Chi_tieu_tuyen_sinh"

    def run(self, dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: Dict[Text, Any]) -> List[Dict[Text, Any]]:

        resultDict = {}
        wb = openpyxl.load_workbook('D:\Projects\RASA\Test-v2.xlsx')
        sheet = wb['Sheet1']
        for i in range(3, 67):
            entity = sheet.cell(row=i, column=6).value
            number = sheet.cell(row=i, column=8).value
            resultDict[entity] = number
        entitiesDict = tracker.latest_message['entities'];
        resultText = "Hiện tại tôi không tìm thấy thông tin liên quan tới ngành bạn quan tâm. Bạn vui lòng kiểm tra lại tên ngành hoặc đặt câu hỏi rõ ràng hơn giúp tôi."
        if len(entitiesDict) > 0 and entity in resultDict:
            entity = entitiesDict[0]["entity"]
            resultText = "Chỉ tiêu xét tuyển của ngành {} là: {}".format(entity, resultDict[entity])
        dispatcher.utter_message(resultText)

        return []

class ActionMaNganh(Action):

    def name(self) -> Text:
        return "action_Ma_nganh"

    def run(self, dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: Dict[Text, Any]) -> List[Dict[Text, Any]]:

        resultDict = {}
        wb = openpyxl.load_workbook('D:\Projects\RASA\Test-v2.xlsx')
        sheet = wb['Sheet1']
        for i in range(3, 67):
            entity = sheet.cell(row=i, column=6).value
            number = sheet.cell(row=i, column=7).value
            resultDict[entity] = number
        entitiesDict = tracker.latest_message['entities'];
        resultText = "Hiện tại tôi không tìm thấy thông tin liên quan tới ngành bạn quan tâm. Bạn vui lòng kiểm tra lại tên ngành hoặc đặt câu hỏi rõ ràng hơn giúp tôi."
        if len(entitiesDict) > 0 and entity in resultDict:
            entity = entitiesDict[0]['entity']
            resultText = "Mã ngành của ngành {} là: {}".format(entity, resultDict[entity])
        dispatcher.utter_message(resultText)

        return []

class ActionDiemXetTuyen(Action):

    def name(self) -> Text:
        return "action_Diem_xet_tuyen"

    def run(self, dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: Dict[Text, Any]) -> List[Dict[Text, Any]]:

        resultDict = {}
        wb = openpyxl.load_workbook('D:\Projects\RASA\Test-v2.xlsx')
        sheet = wb['Sheet1']
        for i in range(3, 67):
            entity = sheet.cell(row=i, column=6).value
            hocBa2020 = sheet.cell(row=i, column=14).value
            hocBa2021 = sheet.cell(row=i, column=16).value
            thpt2020 = sheet.cell(row=i, column=13).value
            thpt2021 = sheet.cell(row=i, column=15).value
            resultDict[entity] = "điểm chuẩn xét tuyển học bạ năm 2021 là: {}; năm 2020 là: {}. Điểm chuẩn THPT năm 2021 là: {}; năm 2020 là: {}".format(hocBa2021, hocBa2020, thpt2021, thpt2020)
        entitiesDict = tracker.latest_message['entities'];
        resultText = "Hiện tại tôi không tìm thấy thông tin liên quan tới ngành bạn quan tâm. Bạn vui lòng kiểm tra lại tên ngành hoặc đặt câu hỏi rõ ràng hơn giúp tôi."
        if len(entitiesDict) > 0 and entity in resultDict:
            entity = entitiesDict[0]['entity']
            resultText = "Ngành {} có {}".format(entity, resultDict[entity])
        dispatcher.utter_message(resultText)

        return []

class ActionToHopXetTuyen(Action):

    def name(self) -> Text:
        return "action_To_hop_xet_tuyen"

    def run(self, dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: Dict[Text, Any]) -> List[Dict[Text, Any]]:

        resultDict = {}
        wb = openpyxl.load_workbook('D:\Projects\RASA\Test-v2.xlsx')
        sheet = wb['Sheet1']
        for i in range(3, 67):
            entity = sheet.cell(row=i, column=6).value
            th1 = sheet.cell(row=i, column=9).value
            th2 = sheet.cell(row=i, column=10).value
            th3 = sheet.cell(row=i, column=11).value
            th4 = sheet.cell(row=i, column=12).value
            resultDict[entity] = "{}, {}, {}, {}".format(th1, th2, th3, th4)
        entitiesDict = tracker.latest_message['entities'];
        resultText = "Hiện tại tôi không tìm thấy thông tin liên quan tới ngành bạn quan tâm. Bạn vui lòng kiểm tra lại tên ngành hoặc đặt câu hỏi rõ ràng hơn giúp tôi."
        if len(entitiesDict) > 0 and entity in resultDict:
            entity = entitiesDict[0]['entity']
            resultText = "Tổ hợp xét tuyển của ngành {} là {}".format(entity, resultDict[entity])
        dispatcher.utter_message(resultText)

        return []

class ActionCongViecRaTruong(Action):

    def name(self) -> Text:
        return "action_Cong_viec_ra_truong"

    def run(self, dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: Dict[Text, Any]) -> List[Dict[Text, Any]]:

        resultDict = {}
        wb = openpyxl.load_workbook('D:\Projects\RASA\Test-v2.xlsx')
        sheet = wb['Sheet1']
        for i in range(3, 67):
            entity = sheet.cell(row=i, column=6).value
            info1 = sheet.cell(row=i, column=18).value
            info2 = sheet.cell(row=i, column=17).value
            resultDict[entity] = str(info1) if len(str(info1))>0 else str(info2)
        entitiesDict = tracker.latest_message['entities'];
        resultText = "Hiện tại tôi không có thông tin liên quan tới ngành bạn quan tâm. Bạn hãy quay lại sau nhé, chúng tôi sẽ bổ sung thông tin."
        if len(entitiesDict) > 0:
            if entity in resultDict and resultDict[entity] != "None":
                entity = entitiesDict[0]['entity']
                resultText = resultDict[entity]
        dispatcher.utter_message(resultText)

        return []

class ActionThongTinNganh(Action):

    def name(self) -> Text:
        return "action_Thong_tin_nganh"

    def run(self, dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: Dict[Text, Any]) -> List[Dict[Text, Any]]:

        resultDict = {}
        wb = openpyxl.load_workbook('D:\Projects\RASA\Test-v2.xlsx')
        sheet = wb['Sheet1']
        for i in range(3, 67):
            entity = sheet.cell(row=i, column=6).value
            info = sheet.cell(row=i, column=17).value
            resultDict[entity] = str(info)
        entitiesDict = tracker.latest_message['entities'];
        print(entitiesDict)
        resultText = "Hiện tại tôi không có thông tin liên quan tới ngành bạn quan tâm"
        if len(entitiesDict) > 0:
            if entity in resultDict and resultDict[entity] != "None":
                entity = entitiesDict[0]['entity']
                resultText = resultDict[entity]
        dispatcher.utter_message(resultText)

        return []

class ActionTiLeViecLam(Action):

    def name(self) -> Text:
        return "action_Ti_le_co_viec_lam"

    def run(self, dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: Dict[Text, Any]) -> List[Dict[Text, Any]]:

        resultDict = {}
        wb = openpyxl.load_workbook('D:\Projects\RASA\Test-v2.xlsx')
        sheet = wb['Sheet1']
        for i in range(3, 67):
            entity = sheet.cell(row=i, column=6).value
            info = sheet.cell(row=i, column=19).value
            resultDict[entity] = str(info)
        entitiesDict = tracker.latest_message['entities'];
        resultText = "Hiện tại tôi không có thông tin về tỉ lệ việc làm của ngành bạn quan tâm vì ngành này chưa có sinh viên ra trường"
        if len(entitiesDict) > 0:
            if entity in resultDict and resultDict[entity] != "Chưa có":
                entity = entitiesDict[0]['entity']
                resultText = "Tỉ lệ việc làm ngành {} là {}".format(entity, resultDict[entity])
        dispatcher.utter_message(resultText)

        return []

class ActionCacNganhDaoTao(Action):

    def name(self) -> Text:
        return "action_Cac_nganh_dao_tao"

    def run(self, dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: Dict[Text, Any]) -> List[Dict[Text, Any]]:

        resultDict = set()
        resultText ="Danh sách các ngành trường tuyển sinh năm 2021 gồm: \n"
        wb = openpyxl.load_workbook('D:\Projects\RASA\Test-v2.xlsx')
        sheet = wb['Sheet1']
        for i in range(3, 67):
            name = sheet.cell(row=i, column=20).value
            resultDict.add(name)
        for item in resultDict:
            resultText += item + '\n'
        dispatcher.utter_message(resultText)

        return []

class ActionCacKhoa(Action):

    def name(self) -> Text:
        return "action_Cac_khoa_cua_truong"

    def run(self, dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: Dict[Text, Any]) -> List[Dict[Text, Any]]:

        resultDict = set()
        resultText ="Danh sách các khoa của trường gồm: \n"
        wb = openpyxl.load_workbook('D:\Projects\RASA\Test-v2.xlsx')
        sheet = wb['Sheet1']
        for i in range(3, 67):
            name = sheet.cell(row=i, column=5).value
            resultDict.add(name)
        for item in resultDict:
            resultText += item + '\n'
        dispatcher.utter_message(resultText)

        return []